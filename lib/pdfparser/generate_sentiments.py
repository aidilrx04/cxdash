import os
import re
import sys
import openpyxl
import pdfplumber
import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from transformers import pipeline

# ==========================================
# ARGUMENT PARSING
# ==========================================
folder = os.path.dirname(__file__)

# Read CLI parameters or default to local fallback files
if len(sys.argv) >= 3:
    filefull = sys.argv[1]
    output_excel = sys.argv[2]
elif len(sys.argv) == 2:
    filefull = sys.argv[1]
    output_excel = os.path.join(folder, "result_analyzed2.xlsx")
else:
    filefull = os.path.join(folder, "sample2.pdf")
    output_excel = os.path.join(folder, "result_analyzed2.xlsx")

# Ensure output directory exists
output_dir = os.path.dirname(output_excel)
if output_dir:
    os.makedirs(output_dir, exist_ok=True)

TARGET_HEADERS = [
    "What did you like most about the course?",
    "What would you change about this course?",
    "The courses you would probably like to attend.",
    "Additional comments",
]

TIMESTAMP_PATTERN = re.compile(
    r"\b\d{1,2}:\d{2}(?::\d{2})?\s*(?:AM|PM|am|pm)?\b"
)
ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# ==========================================
# LOAD AI MODELS
# ==========================================
print("Loading AI models for sentiment analysis and dynamic theme extraction...")

# 1. Zero-shot classifier for sentiment (Positive, Neutral, Negative)
sentiment_classifier = pipeline(
    "zero-shot-classification", model="facebook/bart-large-mnli"
)
SENTIMENT_LABELS = ["positive", "neutral", "negative"]

# 2. Generative text-to-text model for open-ended, question-aware theme extraction
# Note: Swap to "google/flan-t5-base" if you require slightly richer themes on a modern CPU
theme_generator = pipeline(
    "text2text-generation", model="google/flan-t5-small"
)


def clean_cell(cell):
    """Clean newlines, outer whitespace, timestamps, and illegal Excel characters."""
    if cell is None:
        return ""
    text = str(cell).replace("\n", " ").strip()
    text = TIMESTAMP_PATTERN.sub("", text).strip()
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text


def analyze_text(text, header=""):
    """Detect sentiment using zero-shot classification and generate a dynamic theme using FLAN-T5."""
    if not text or len(text.strip()) < 2:
        return "", ""

    # 1. Sentiment classification
    sentiment_res = sentiment_classifier(
        text, candidate_labels=SENTIMENT_LABELS, multi_label=False
    )
    top_sentiment = sentiment_res["labels"][0]

    # 2. Open-ended dynamic theme extraction with Question Context
    if header:
        prompt = (
            f"Question: {header}\n"
            f"Answer: {text}\n"
            f"Summarize the main topic or feedback of the answer in 2 to 4 words:"
        )
    else:
        prompt = f"Summarize the main topic of this feedback in 2 to 4 words: {text}"

    theme_res = theme_generator(prompt, max_new_tokens=12, do_sample=False)
    top_theme = theme_res[0]["generated_text"].strip().strip(".")

    return top_sentiment, top_theme


def identify_header(row):
    """Identify which target question header is present in the row."""
    row_text = " ".join(
        [str(col).replace("\n", " ").strip() for col in row if col]
    ).lower()
    for target in TARGET_HEADERS:
        if target.lower() in row_text:
            return target
    return None


# ==========================================
# STEP 1: EXTRACT & MERGE DATA FROM PDF
# ==========================================
merged_people = {}
current_active_header = None

with pdfplumber.open(filefull) as pdf:
    for page in pdf.pages:
        tables = page.extract_tables()

        for table in tables:
            if not table:
                continue

            detected_header = identify_header(table[0])

            if detected_header:
                current_active_header = detected_header
                rows_to_process = table[1:]
            elif current_active_header:
                rows_to_process = table
            else:
                continue

            last_name = None
            for row in rows_to_process:
                cleaned_row = [clean_cell(cell) for cell in row]

                if len(cleaned_row) < 3:
                    continue

                no_val = cleaned_row[0]
                name_val = cleaned_row[1]
                response_val = cleaned_row[2]

                if no_val.isdigit() and name_val:
                    last_name = name_val
                    if last_name not in merged_people:
                        merged_people[last_name] = {}
                    merged_people[last_name][
                        current_active_header
                    ] = response_val

                elif last_name and response_val:
                    existing_ans = merged_people[last_name].get(
                        current_active_header, ""
                    )
                    merged_people[last_name][current_active_header] = (
                        f"{existing_ans} {response_val}".strip()
                    )

# ==========================================
# STEP 2: RUN AI ANALYSIS & BUILD ROWS
# ==========================================
final_columns = ["No", "Name"]
for header in TARGET_HEADERS:
    final_columns.extend(
        [
            header,
            f"sentiment_{header}",
            f"theme_{header}",
            f"correctness_{header}",
        ]
    )

rows_data = []
print("Running sentiment analysis and dynamic theme extraction...")

for idx, (name, responses) in enumerate(merged_people.items(), start=1):
    person_row = {"No": idx, "Name": name}

    for header in TARGET_HEADERS:
        response_text = responses.get(header, "")
        person_row[header] = response_text

        # Pass both response text and question header into analyze_text
        sentiment, theme = analyze_text(response_text, header=header)

        person_row[f"sentiment_{header}"] = sentiment
        person_row[f"theme_{header}"] = theme
        person_row[f"correctness_{header}"] = ""  # Left blank for manual review

    rows_data.append(person_row)

df = pd.DataFrame(rows_data, columns=final_columns)

display_columns = ["No", "Name"]
for header in TARGET_HEADERS:
    display_columns.extend([header, "sentiment", "theme", "correctness"])
df.columns = display_columns

# ==========================================
# STEP 3: EXPORT TO EXCEL WITH DROPDOWNS & COLORS
# ==========================================
with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Analyzed Feedback")

# Re-open workbook via openpyxl to apply Excel Data Validation and Styling
wb = openpyxl.load_workbook(output_excel)
ws = wb["Analyzed Feedback"]

# Data Validation Rule (Dropdown List)
dv = DataValidation(
    type="list", formula1='"Correct,Incorrect"', allow_blank=True
)
ws.add_data_validation(dv)

# Styling for Conditional Formatting
green_fill = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)
green_font = Font(color="006100")
red_fill = PatternFill(
    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
)
red_font = Font(color="9C0006")

rule_correct = CellIsRule(
    operator="equal", formula=['"Correct"'], fill=green_fill, font=green_font
)
rule_incorrect = CellIsRule(
    operator="equal", formula=['"Incorrect"'], fill=red_fill, font=red_font
)

max_row = len(df) + 1

# Identify all "correctness" columns and attach validation + conditional formatting
for col_idx, col_name in enumerate(df.columns, start=1):
    if col_name == "correctness":
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{max_row}"

        # Attach Dropdown
        dv.add(cell_range)

        # Attach Color Rules
        ws.conditional_formatting.add(cell_range, rule_correct)
        ws.conditional_formatting.add(cell_range, rule_incorrect)

# Auto-adjust Column Widths
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    col_letter = col[0].column_letter
    ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

wb.save(output_excel)
print(
    f"Successfully exported Excel file with dropdowns and formatting to: {output_excel}"
)